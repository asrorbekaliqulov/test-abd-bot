import asyncio
import random
import os
from datetime import datetime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
import openpyxl
import requests
from django.conf import settings
from asgiref.sync import sync_to_async

# Django modelni import qilish
from apps.Bot.models.TelegramBot import TelegramUser  # O'z app nomingizni kiriting


# States
SELECTING_TEST, SELECTING_CATEGORY, WAITING_EXCEL, CREATE_TEST_NAME = range(4)

# Backend URL va Log kanal
BACKEND_URL = settings.BASE_API_URL if hasattr(settings, 'BASE_API_URL') else "https://backend.testabd.uz/"
LOG_CHANNEL_ID = settings.LOG_CHANNEL_ID if hasattr(settings, 'LOG_CHANNEL_ID') else -1003642436546

@sync_to_async
def get_user_access_token(user_id: int) -> str:
    """Django modeldan access token olish"""
    try:
        user = TelegramUser.objects.get(user_id=user_id)
        print(f"Foydalanuvchi topildi: {user_id}")
        return user.access_token
    except TelegramUser.DoesNotExist:
        return None
    except Exception as e:
        return None


def get_headers(access_token: str) -> dict:
    """API uchun headers"""
    if not access_token:
        return {}
    return {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }


async def start_create_question(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savol yaratishni boshlash - /create_question"""
    user_id = update.effective_user.id
    access_token = await get_user_access_token(user_id)
    
    if not access_token:
        await update.message.reply_text(
            "❌ Avval botdan ro'yxatdan o'ting!\n"
            "Ro'yxatdan o'tish uchun /start buyrug'ini yuboring."
        )
        return ConversationHandler.END
    
    # Foydalanuvchi testlarini olish
    try:
        response = requests.get(
            f"{BACKEND_URL}/quiz/tests/my_tests/",
            headers=get_headers(access_token),
            timeout=10
        )
        response.raise_for_status()
        tests = response.json()
        
        if not tests:
            await update.message.reply_text(
                "📚 Sizda hali testlar yo'q.\n\n"
                "Yangi test yaratish uchun quyidagi tugmani bosing:"
            )
            keyboard = [[InlineKeyboardButton("➕ Yangi test yaratish", callback_data="create_new_test")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text("Test yarating:", reply_markup=reply_markup)
        else:
            # Inline keyboard yaratish
            keyboard = []
            for test in tests:
                question_count = test.get('total_questions', '0')
                # String yoki int bo'lishi mumkin
                try:
                    q_count = int(question_count)
                except:
                    q_count = 0
                    
                emoji = "🔵" if q_count == 0 else "✅"
                button_text = f"{emoji} {test['title']} ({q_count} ta savol)"
                keyboard.append([InlineKeyboardButton(
                    button_text,
                    callback_data=f"test_{test['id']}"
                )])
            
            # Yangi test yaratish tugmasi
            keyboard.append([InlineKeyboardButton(
                "➕ Yangi test yaratish",
                callback_data="create_new_test"
            )])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            await update.message.reply_text(
                "📚 Sizning testlaringiz:\n\nTest tanlang yoki yangi yarating:",
                reply_markup=reply_markup
            )
        
        context.user_data['access_token'] = access_token
        return SELECTING_TEST
        
    except requests.RequestException as e:
        print(f"Testlarni olishda xatolik: {e}")
        await update.message.reply_text(
            "❌ Testlarni yuklashda xatolik yuz berdi!\n"
            "Iltimos, qaytadan urinib ko'ring."
        )
        return ConversationHandler.END


async def handle_test_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test tanlash"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "create_new_test":
        await query.edit_message_text("📝 Yangi test nomini kiriting:")
        return CREATE_TEST_NAME
    
    # Test ID ni olish
    test_id = query.data.replace("test_", "")
    context.user_data['selected_test_id'] = test_id
    
    # Kategoriyalarni ko'rsatish
    return await show_categories(query, context)


async def create_new_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Yangi test yaratish"""
    test_name = update.message.text.strip()
    
    if len(test_name) < 3:
        await update.message.reply_text(
            "❌ Test nomi kamida 3 ta belgidan iborat bo'lishi kerak!\n"
            "Iltimos, qaytadan kiriting:"
        )
        return CREATE_TEST_NAME
    
    context.user_data['new_test_name'] = test_name
    
    # Kategoriyalarni olish
    access_token = context.user_data['access_token']
    
    try:
        response = requests.get(
            f"{BACKEND_URL}/quiz/categories",
            headers=get_headers(access_token),
            timeout=10
        )
        response.raise_for_status()
        categories = response.json()
        
        # 3 ta ustunda inline keyboard
        keyboard = []
        row = []
        for i, cat in enumerate(categories):
            emoji = cat.get('emoji', '📁')
            button = InlineKeyboardButton(
                f"{emoji} {cat['title']}",
                callback_data=f"newcat_{cat['id']}"
            )
            row.append(button)
            
            if len(row) == 3 or i == len(categories) - 1:
                keyboard.append(row)
                row = []
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "📂 Test uchun kategoriya tanlang:",
            reply_markup=reply_markup
        )
        return SELECTING_CATEGORY
        
    except requests.RequestException as e:
        print(f"Kategoriyalarni olishda xatolik: {e}")
        await update.message.reply_text("❌ Kategoriyalarni yuklashda xatolik!")
        return ConversationHandler.END


async def show_categories(query, context: ContextTypes.DEFAULT_TYPE):
    """Kategoriyalarni ko'rsatish"""
    access_token = context.user_data['access_token']
    
    try:
        response = requests.get(
            f"{BACKEND_URL}/quiz/categories",
            headers=get_headers(access_token),
            timeout=10
        )
        response.raise_for_status()
        categories = response.json()
        
        # 3 ta ustunda keyboard
        keyboard = []
        row = []
        for i, cat in enumerate(categories):
            emoji = cat.get('emoji', '📁')
            button = InlineKeyboardButton(
                f"{emoji} {cat['title']}",
                callback_data=f"cat_{cat['id']}"
            )
            row.append(button)
            
            if len(row) == 3 or i == len(categories) - 1:
                keyboard.append(row)
                row = []
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "📂 Savol qo'shish uchun kategoriya tanlang:",
            reply_markup=reply_markup
        )
        return SELECTING_CATEGORY
        
    except requests.RequestException as e:
        print(f"Kategoriyalarni olishda xatolik: {e}")
        await query.edit_message_text("❌ Kategoriyalarni yuklashda xatolik!")
        return ConversationHandler.END


async def handle_category_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kategoriya tanlash"""
    query = update.callback_query
    await query.answer()
    
    # Yangi test yaratish
    if query.data.startswith("newcat_"):
        category_id = query.data.replace("newcat_", "")
        test_name = context.user_data['new_test_name']
        access_token = context.user_data['access_token']
        
        # Test yaratish
        test_data = {
            "title": test_name,
            "category_id": int(category_id),
            "visibility": "public",
            "access_mode": "open",
            "participant_roles": "users_only"
        }
        
        try:
            response = requests.post(
                f"{BACKEND_URL}/quiz/tests",
                headers=get_headers(access_token),
                json=test_data,
                timeout=10
            )
            print(f"Test yaratish javobi: {response.status_code} - {response.text}")
            response.raise_for_status()
            test = response.json()
            context.user_data['selected_test_id'] = test['id']
            context.user_data['category_id'] = category_id
            
            await query.edit_message_text(
                f"✅ Test '{test_name}' muvaffaqiyatli yaratildi!\n\n"
                "📎 Endi Excel fayllarni yuboring (bir yoki bir nechta fayl).\n\n"
                "📋 Excel shablon:\n"
                "• 1-ustun: Savol matni\n"
                "• 2-ustun: To'g'ri javob\n"
                "• 3-ustun: Noto'g'ri javob 1\n"
                "• 4-ustun: Noto'g'ri javob 2\n\n"
                "✅ Fayllarni yuborish tugagach /done buyrug'ini yuboring."
            )
            context.user_data['excel_files'] = []
            return WAITING_EXCEL
            
        except requests.RequestException as e:
            print(f"Test yaratishda xatolik: {e}")
            await query.edit_message_text("❌ Test yaratishda xatolik yuz berdi!")
            return ConversationHandler.END
    
    # Mavjud testga savol qo'shish
    else:
        category_id = query.data.replace("cat_", "")
        context.user_data['category_id'] = category_id
        
        await query.edit_message_text(
            "📎 Excel fayllarni yuboring (bir yoki bir nechta fayl).\n\n"
            "📋 Excel shablon:\n"
            "• 1-ustun: Savol matni\n"
            "• 2-ustun: To'g'ri javob\n"
            "• 3-ustun: Noto'g'ri javob 1\n"
            "• 4-ustun: Noto'g'ri javob 2\n\n"
            "✅ Fayllarni yuborish tugagach /done buyrug'ini yuboring."
        )
        context.user_data['excel_files'] = []
        return WAITING_EXCEL


async def receive_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Excel fayllarni qabul qilish"""
    if update.message.document:
        file = update.message.document
        
        if file.file_name.endswith(('.xlsx', '.xls')):
            try:
                # Faylni yuklab olish
                file_obj = await context.bot.get_file(file.file_id)
                file_path = f"temp_{file.file_id}.xlsx"
                await file_obj.download_to_drive(file_path)
                
                if 'excel_files' not in context.user_data:
                    context.user_data['excel_files'] = []
                    
                context.user_data['excel_files'].append(file_path)
                
                file_count = len(context.user_data['excel_files'])
                await update.message.reply_text(
                    f"✅ Fayl qabul qilindi: {file.file_name}\n"
                    f"📊 Jami fayllar: {file_count}\n\n"
                    "Yana fayl yuborishingiz yoki /done buyrug'ini yuborishingiz mumkin."
                )
            except Exception as e:
                print(f"Faylni yuklashda xatolik: {e}")
                await update.message.reply_text("❌ Faylni yuklashda xatolik!")
        else:
            await update.message.reply_text("❌ Faqat Excel fayl (.xlsx, .xls) qabul qilinadi!")
    
    return WAITING_EXCEL


async def process_excel_files(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Excel fayllarni qayta ishlash"""
    excel_files = context.user_data.get('excel_files', [])
    
    if not excel_files:
        await update.message.reply_text("❌ Hech qanday Excel fayl yuklanmagan!")
        return WAITING_EXCEL
    
    await update.message.reply_text(
        f"⏳ Savollar yaratilmoqda...\n"
        f"📊 Jami {len(excel_files)} ta fayl\n\n"
        "Bu jarayon biroz vaqt olishi mumkin. Iltimos, kuting..."
    )
    
    test_id = context.user_data['selected_test_id']
    category_id = context.user_data.get('category_id')
    access_token = context.user_data['access_token']
    
    total_success = 0
    total_failed = 0
    failed_questions = []
    
    # Har bir Excel fayl uchun
    for file_index, file_path in enumerate(excel_files):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Progress xabari
            await update.message.reply_text(
                f"📄 {file_index + 1}/{len(excel_files)} - Fayl ishlanmoqda..."
            )
            
            # Savollarni o'qish (2-qatordan boshlab, 1-qator header)
            for row_num in range(2, sheet.max_row + 1):
                try:
                    question_text = sheet.cell(row_num, 1).value
                    correct_answer = sheet.cell(row_num, 2).value
                    wrong_answer1 = sheet.cell(row_num, 3).value
                    wrong_answer2 = sheet.cell(row_num, 4).value
                    
                    # Bo'sh qatorlarni o'tkazib yuborish
                    if not all([question_text, correct_answer, wrong_answer1, wrong_answer2]):
                        continue
                    
                    # Variantlarni tayyorlash
                    answers = [
                        {"answer_text": str(correct_answer).strip(), "is_correct": True},
                        {"answer_text": str(wrong_answer1).strip(), "is_correct": False},
                        {"answer_text": str(wrong_answer2).strip(), "is_correct": False}
                    ]
                    
                    # Random aralashtirish
                    random.shuffle(answers)
                    
                    # Harflarni qo'shish (A, B, D)
                    letters = ['A', 'B', 'D']
                    for i, ans in enumerate(answers):
                        ans["letter"] = letters[i]
                    # Savol yaratish
                    question_data = {
                        "test": int(test_id),
                        "question_text": str(question_text).strip(),
                        "question_type": "single",
                        "order_index": total_success + total_failed,
                        "answers": answers,
                        "category_id": int(category_id) if category_id else None
                    }
                    
                    response = requests.post(
                        f"{BACKEND_URL}/quiz/questions/",
                        headers=get_headers(access_token),
                        json=question_data,
                        timeout=15
                    )
                    print(f"Response status: {response.status_code}, Question: {response.text}...")
                    if response.status_code in [200, 201]:
                        total_success += 1
                    else:
                        total_failed += 1
                        error_msg = response.json().get('detail', 'Noma\'lum xatolik')
                        failed_questions.append(
                            f"Fayl {file_index + 1}, Qator {row_num}: {str(question_text)[:40]}... - {error_msg}"
                        )
                    
                    # Har bir savol orasida 5-7 soniya kutish
                    await asyncio.sleep(random.uniform(5, 7))
                    
                except Exception as e:
                    print(f"Savol yaratishda xatolik (qator {row_num}): {e}")
                    total_failed += 1
                    failed_questions.append(
                        f"Fayl {file_index + 1}, Qator {row_num}: Xatolik - {str(e)}"
                    )
            
            # Faylni o'chirish
            try:
                os.remove(file_path)
            except:
                pass
                
        except Exception as e:
            print(f"Excel faylni o'qishda xatolik: {e}")
            await update.message.reply_text(f"❌ {file_path} faylni o'qishda xatolik!")
    
    # Natijani yuborish
    result_text = (
        f"📊 <b>Natijalar:</b>\n\n"
        f"✅ Muvaffaqiyatli: <b>{total_success}</b>\n"
        f"❌ Xatolik: <b>{total_failed}</b>\n"
    )
    
    if failed_questions and len(failed_questions) <= 10:
        result_text += "\n❌ <b>Muvaffaqiyatsiz savollar:</b>\n"
        for i, failed in enumerate(failed_questions, 1):
            result_text += f"{i}. {failed}\n"
    elif failed_questions:
        result_text += f"\n❌ <b>Muvaffaqiyatsiz savollar:</b> {len(failed_questions)} ta\n"
        result_text += "Birinchi 5 tasi:\n"
        for i, failed in enumerate(failed_questions[:5], 1):
            result_text += f"{i}. {failed}\n"
    
    await update.message.reply_text(result_text, parse_mode='HTML')
    
    # Log kanalga yuborish
    if LOG_CHANNEL_ID:
        try:
            user = update.effective_user
            log_text = (
                f"📝 <b>Yangi savollar qo'shildi</b>\n\n"
                f"👤 Foydalanuvchi: @{user.username or 'username_yoq'} ({user.id})\n"
                f"🆔 Test ID: {test_id}\n"
                f"📁 Fayllar soni: {len(excel_files)}\n"
                f"✅ Muvaffaqiyatli: {total_success}\n"
                f"❌ Xatolik: {total_failed}\n"
                f"📅 Sana: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            await context.bot.send_message(LOG_CHANNEL_ID, log_text, parse_mode='HTML')
        except Exception as e:
            print(f"Log kanalga yuborishda xatolik: {e}")
    
    # User data tozalash
    context.user_data.clear()
    
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bekor qilish"""
    # Vaqtinchalik fayllarni o'chirish
    excel_files = context.user_data.get('excel_files', [])
    for file_path in excel_files:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except:
            pass
    
    context.user_data.clear()
    await update.message.reply_text("❌ Amal bekor qilindi!")
    return ConversationHandler.END


# ConversationHandler yaratish
def get_create_question_handler():
    """ConversationHandler qaytarish - bu funksiyani botga qo'shing"""
    return ConversationHandler(
        entry_points=[CommandHandler("create_question", start_create_question)],
        states={
            SELECTING_TEST: [
                CallbackQueryHandler(handle_test_selection)
            ],
            CREATE_TEST_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, create_new_test)
            ],
            SELECTING_CATEGORY: [
                CallbackQueryHandler(handle_category_selection)
            ],
            WAITING_EXCEL: [
                MessageHandler(filters.Document.ALL, receive_excel),
                CommandHandler("done", process_excel_files)
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel)
        ],
    )


# Botga qo'shish uchun (asosiy bot faylingizda):
# from your_handlers_file import get_create_question_handler
# application.add_handler(get_create_question_handler())